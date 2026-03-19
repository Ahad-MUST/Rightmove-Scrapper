"""
Rightmove Scraper - Data Saver
================================
Handles JSON, CSV, and styled Excel output.
"""

import json
import re
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config


# ── Column definitions ────────────────────────────────────────────────────────

PROP_COLUMNS = [
    'property_id', 'price', 'agent_phone', 'url',
    'description', 'key_features',
    'address', 'property_type', 'bedrooms',
    'search_area', 'search_city', 'filter_label', 'scraped_at',
]

COL_WIDTHS = {
    'property_id':  14,
    'price':        14,
    'agent_phone':  18,
    'url':          45,
    'description':  55,
    'key_features': 42,
    'address':      35,
    'property_type':16,
    'bedrooms':     10,
    'search_area':  22,
    'search_city':  14,
    'filter_label': 36,
    'scraped_at':   18,
}

# Colour palette
COLOR_AREA_HEADER   = 'FF2C3E50'  # dark navy  – area row
COLOR_FILTER_HEADER = 'FF1ABC9C'  # teal       – filter section
COLOR_COL_HEADER    = 'FF2980B9'  # blue       – column headers
COLOR_ROW_ALT       = 'FFF0F8FF'  # light blue – alternate data rows
COLOR_SUMMARY_BG    = 'FFFEF9E7'  # pale yellow – summary sheet
COLOR_SUMMARY_HDR   = 'FFF39C12'  # orange     – summary header


# ── Small utilities ───────────────────────────────────────────────────────────

def _bed_display(val) -> str:
    """Convert a bedroom value to a display string. 0 → 'Studio'."""
    if val is None or val == '':
        return ''
    try:
        return 'Studio' if int(val) == 0 else str(int(val))
    except (TypeError, ValueError):
        return str(val)


def _thin_border() -> Border:
    s = Side(style='thin', color='FFB0BEC5')
    return Border(left=s, right=s, top=s, bottom=s)


# ── Excel row writers ─────────────────────────────────────────────────────────

def _write_area_header(ws, row: int, area_name: str, num_cols: int):
    cell = ws.cell(row=row, column=1, value=f"  {area_name}")
    cell.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=13)
    cell.fill      = PatternFill('solid', start_color=COLOR_AREA_HEADER)
    cell.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[row].height = 22
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', start_color=COLOR_AREA_HEADER)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)


def _write_filter_header(ws, row: int, filter_label: str, num_cols: int):
    cell = ws.cell(row=row, column=1, value=f"   {filter_label}")
    cell.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=11)
    cell.fill      = PatternFill('solid', start_color=COLOR_FILTER_HEADER)
    cell.alignment = Alignment(vertical='center', indent=2)
    ws.row_dimensions[row].height = 20
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', start_color=COLOR_FILTER_HEADER)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)


def _write_col_headers(ws, row: int, cols: List[str]):
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=col.replace('_', ' ').title())
        cell.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
        cell.fill      = PatternFill('solid', start_color=COLOR_COL_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[row].height = 18


def _write_property_row(ws, row: int, prop: dict, cols: List[str], alt: bool):
    bg = COLOR_ROW_ALT if alt else 'FFFFFFFF'
    for ci, col in enumerate(cols, 1):
        val = prop.get(col, '')
        if isinstance(val, list):
            val = ' | '.join(str(v) for v in val)
        if col == 'bedrooms':
            val = _bed_display(val)
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font      = Font(name='Arial', size=9)
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(
            wrap_text=(col in ('description', 'key_features', 'address')),
            vertical='top',
        )
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 15


def _set_col_widths(ws, cols: List[str]):
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)


# ── Public API ────────────────────────────────────────────────────────────────

class DataSaver:

    @staticmethod
    def save_to_json(data: List[Dict], filename: str):
        """Write a flat list of property dicts to <filename>.json."""
        filepath = f"{filename}.json"
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=config.JSON_INDENT, ensure_ascii=False)
        print(f"Saved JSON: {filepath}")

    @staticmethod
    def save_to_csv(data: List[Dict], filename: str):
        """Write a flat list of property dicts to <filename>.csv."""
        if not data:
            return
        flat = []
        for item in data:
            fi = item.copy()
            for k in ('images', 'key_features'):
                if isinstance(fi.get(k), list):
                    fi[k] = ', '.join(fi[k])
            if 'bedrooms' in fi:
                fi['bedrooms'] = _bed_display(fi['bedrooms'])
            flat.append(fi)
        pd.DataFrame(flat).to_csv(f"{filename}.csv", index=False, encoding='utf-8')
        print(f"Saved CSV: {filename}.csv")

    @staticmethod
    def save_organized_excel(organized: Dict, filepath: str):
        """
        Save scrape results to a styled Excel workbook.

        Structure
        ---------
        - One sheet per area, with one colour-coded section per filter set.
        - A Summary sheet with property counts, average prices, and bed ranges.
        """
        wb = Workbook()
        wb.remove(wb.active)

        cols     = PROP_COLUMNS
        num_cols = len(cols)

        # ── Summary sheet ──────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        ws_sum.sheet_properties.tabColor = "F39C12"

        title_cell = ws_sum.cell(row=1, column=1, value="Rightmove Scrape – Summary")
        title_cell.font = Font(name='Arial', bold=True, size=14, color='FF2C3E50')
        ws_sum.merge_cells('A1:F1')
        ws_sum.row_dimensions[1].height = 24

        for ci, header in enumerate(
            ['Area', 'Filter Set', 'Properties Found', 'Min Price', 'Avg Price', 'Max Price', 'Min Beds', 'Max Beds'], 1
        ):
            c = ws_sum.cell(row=3, column=ci, value=header)
            c.font      = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
            c.fill      = PatternFill('solid', start_color=COLOR_SUMMARY_HDR)
            c.alignment = Alignment(horizontal='center')
            c.border    = _thin_border()
        ws_sum.row_dimensions[3].height = 18

        for col_letter, width in zip('ABCDEFGH', [28, 40, 18, 16, 16, 16, 12, 12]):
            ws_sum.column_dimensions[col_letter].width = width

        sum_row = 4

        # ── One sheet per area ─────────────────────────────────────────────
        for area_name, filter_data in organized.items():
            ws          = wb.create_sheet(area_name[:31])
            current_row = 1

            _write_area_header(ws, current_row, area_name, num_cols)
            current_row += 1

            for filter_label, props in filter_data.items():
                ws.row_dimensions[current_row].height = 8
                current_row += 1

                _write_filter_header(ws, current_row, filter_label, num_cols)
                current_row += 1

                _write_col_headers(ws, current_row, cols)
                current_row += 1

                if props:
                    for pi, prop in enumerate(props):
                        _write_property_row(ws, current_row, prop, cols, pi % 2 == 1)
                        current_row += 1
                else:
                    cell = ws.cell(
                        row=current_row, column=1,
                        value="  No properties found for this filter set",
                    )
                    cell.font = Font(name='Arial', italic=True, color='FF999999', size=9)
                    ws.merge_cells(
                        start_row=current_row, start_column=1,
                        end_row=current_row,   end_column=num_cols,
                    )
                    current_row += 1

                # Summary stats
                beds_vals = []
                for p in props:
                    b = p.get('bedrooms')
                    if b is not None and b != '':
                        try:
                            beds_vals.append(int(b))
                        except (TypeError, ValueError):
                            pass

                prices_num = []
                for p in props:
                    cleaned = re.sub(r'[^\d]', '', str(p.get('price') or ''))
                    if cleaned:
                        try:
                            prices_num.append(int(cleaned))
                        except ValueError:
                            pass

                avg_price        = f"£{int(sum(prices_num) / len(prices_num)):,}" if prices_num else "—"
                min_price        = f"£{min(prices_num):,}" if prices_num else "—"
                max_price        = f"£{max(prices_num):,}" if prices_num else "—"
                min_beds_display = _bed_display(min(beds_vals)) if beds_vals else "—"
                max_beds_display = _bed_display(max(beds_vals)) if beds_vals else "—"

                for ci, value in enumerate(
                    [area_name, filter_label, len(props), min_price, avg_price,
                     max_price, min_beds_display, max_beds_display], 1
                ):
                    c = ws_sum.cell(row=sum_row, column=ci, value=value)
                    c.font      = Font(name='Arial', size=9)
                    c.fill      = PatternFill('solid', start_color=COLOR_SUMMARY_BG)
                    c.alignment = Alignment(horizontal='center' if ci > 2 else 'left')
                    c.border    = _thin_border()

                sum_row += 1

            _set_col_widths(ws, cols)
            ws.freeze_panes = 'A2'

        wb.save(filepath)
        print(f"Saved organized Excel: {filepath}")

    @staticmethod
    def print_stats(properties: List[Dict]):
        """Print a quick extraction-rate summary to stdout."""
        if not properties:
            print("No properties to analyse")
            return
        total = len(properties)
        for field, label in [
            ('price',       'Prices'),
            ('address',     'Addresses'),
            ('agent_phone', 'Agent phones'),
            ('description', 'Descriptions'),
        ]:
            count = sum(1 for p in properties if p.get(field))
            print(f"  {label}: {count}/{total} ({count / total * 100:.0f}%)")